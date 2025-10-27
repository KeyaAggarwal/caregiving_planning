import openpyxl

wb = openpyxl.load_workbook("annotations/encoding/coding manual_UPDATED.xlsx")
sheet = wb.active

# Try reading each sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        print(row)
        if i > 10:
            break
